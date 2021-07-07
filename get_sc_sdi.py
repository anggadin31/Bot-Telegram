from pymongo import MongoClient
from bson.objectid import ObjectId
from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
import os
from pymongo import MongoClient, InsertOne
from collections import defaultdict
from selenium import webdriver
import time
from selenium.webdriver.firefox.options import Options

client = MongoClient()
db = client.unsc
sditomman = db.sditomman
tanggal = datetime.now().strftime('%Y%m%d')




link_login = 'http://sdi.tomman.info:8041/login'
link_dl = 'http://sdi.tomman.info:8041/list/all/all/total/export/all/all'
print(link_dl)
options = Options();
options.set_preference("browser.download.folderList",2);
options.set_preference("browser.download.manager.showWhenStarting", False);
options.set_preference("browser.download.dir","D:\damanbot");
options.set_preference("browser.helperApps.neverAsk.saveToDisk", 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet');
driver = webdriver.Firefox(firefox_options=options);

result = None
while result is None:
    try:
        print('in')
        driver.get(link_login)
        print "loading..."
        uname = 'winda  '
        pw = 'TelkomAMO'
        username = driver.find_element_by_name("user")
        password = driver.find_element_by_name("pass")
        username.send_keys(uname)
        password.send_keys(pw)
        login = driver.find_element_by_xpath("//button[@type='submit']")
        login.click()
        time.sleep(10)
        result = driver.find_element_by_xpath("//a[@href='/logout']")
    except Exception as e:
        print e
        print "login failed"
operations = []
try:
    os.remove(os.getcwd()+'\New file.xlsx') 
except:
    pass
driver.set_page_load_timeout(100)
try:
    driver.get(link_dl)
except:
    pass
print('get data')
wb = load_workbook(filename = os.getcwd()+'/New file.xlsx', read_only=True, data_only=True)
ws = wb.active
data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, min_row=2)]

for x in data:
    
    toinsert = {
                'sc': x[1],
                'status' : x[2],
                'koorsc' : x[3],
                'koorplg': x[4],
                'loker' : x[5],
                'progressun' : x[6],
                'tanggal' : tanggal
                }
    operations.append(InsertOne(toinsert))
print('selesai')
sditomman.drop()
sditomman.bulk_write(operations)
print('tutup')
driver.close()

