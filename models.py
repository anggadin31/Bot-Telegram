# from pymongo import MongoClient
import datetime
from openpyxl import load_workbook
from pymongo import MongoClient, InsertOne
from collections import defaultdict
import gspread
from oauth2client.service_account import ServiceAccountCredentials


def nestdict():
    return defaultdict(nestdict)        

class Excelreport():

    def __init__(self):
        client = MongoClient()
        self.db = client.daman
        self.coll = self.db.excelreport

    def get_started(self):      
        operations = []
        data = nestdict()
        pos = 0
        path='D:\daman\EXCEL REPORT_ODP.xlsx'
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb['OCC']    
        data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, max_col=16, min_row=2)]
        print (data[0])

        for x in data:
            try:
                odp = data[data.index(x)][2].split(' ')
            except:
                pass
            toinsert = {
                        'LON': x[0],
                        'LAT' : x[1],
                        'ODP' : odp[0],
                        'RESULT ODP': x[3],
                        'STO' : x[4],
                        'KAP' : x[5],
                        'USED' : x[6],
                        'AVAI' : x[7],
                        'PORT OLT' : x[8],
                        'HOSTNAME' : x[9],
                        'OCC' : x[10],
                        'TANGGAL R2C' : x[11],
                        'BULAN R2C' : x[12],
                        'MITRA' : x[13],
                        'KET' : x[14],
                        'STATUS SIIS' : x[15],
                        }
            operations.append(InsertOne(toinsert))
            pos +=1
        self.coll.drop()
        result = self.coll.bulk_write(operations)
        return pos

    def get_odp(self):
        data = self.coll.find()
        return [each['ODP'] for each in data]

    def get_data(self, odp):
        data = self.coll.find_one({'ODP':{'$regex':odp}})
        return data

class IDPort():

    def __init__(self):
        client = MongoClient()
        self.db = client.daman
        self.coll = self.db.idport

    def get_started(self):      
        operations = []
        data = nestdict()
        pos = 0
        path='D:\daman\ID Port OLT Banjarmasin (new update).xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb['Data']    
        data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, max_col=6, min_row=2)]

        for x in data:
            toinsert = {
                        'IP': x[0],
                        'SLOT' : x[1],
                        'PORT' : x[2],
                        'IP SLOT PORT': x[3],
                        'ID PORT' : x[4],
                        'ID LOGICAL DEVICE' : x[5],
                        }
            operations.append(InsertOne(toinsert))
            pos +=1
        self.coll.drop()
        result = self.coll.bulk_write(operations)
        return pos

    def get_ipslotport(self):
        data = self.coll.find()
        return [each['IP SLOT PORT'] for each in data]

    def get_data(self, ipslotport):
        data = self.coll.find_one({'IP SLOT PORT':{'$regex':ipslotport}})
        return data

class Golive():

    def __init__(self):
        client = MongoClient()
        self.db = client.daman
        self.coll = self.db.golive

    def get_started(self):
        
        operations = []
        data = nestdict()
        pos = 0
        path='D:\daman\CEK GOLIVE.xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb['Sheet2']    
        data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, max_col=19, min_row=2)]
        print (data[0])

        for x in data:
            toinsert = {
                        'ODP_NAME': x[2],
                        'LAT' : x[3],
                        'LON' : x[4],
                        'DATEL': x[14],
                        'STO' : x[15],
                        }
            operations.append(InsertOne(toinsert))
            pos +=1
        self.coll.drop()
        result = self.coll.bulk_write(operations)
        return pos

    def get_odp(self):
        data = self.coll.find()
        return [each['ODP_NAME'] for each in data]

    def get_data(self, odp):
        data = self.coll.find_one({'ODP_NAME':{'$regex':odp}})
        return data

class QrOdp():

    def __init__(self):
        client = MongoClient()
        self.db = client.daman
        self.coll = self.db.qrodp

    def get_started(self):
        
        operations = []
        data = nestdict()
        pos = 0
        path='D:\daman\DATA_QR_ODP.xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb['validasi_odp_witel_KALSEL']    
        data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=6, max_col=8, min_row=2)]
        print (data[0])

        for x in data:
            toinsert = {
                        'ODP_NAME': x[0],
                        'QR CODE' : x[2],
                        }
            operations.append(InsertOne(toinsert))
            pos +=1
        self.coll.drop()
        result = self.coll.bulk_write(operations)
        return pos

    def get_odp(self):
        data = self.coll.find()
        return [each['ODP_NAME'] for each in data]

    def get_data(self, odp):
        data = self.coll.find_one({'ODP_NAME':{'$regex':odp}})
        return data
        
class OdpUim():

    def __init__(self):
        client = MongoClient()
        self.db = client.daman
        self.coll = self.db.odpuim

    def get_started(self):
        
        operations = []
        data = nestdict()
        pos = 0
        path='D:\daman\DALAPA_VALIDASI_UIM.xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb['Sheet2']    
        data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, max_col=15, min_row=2)]
        print(data[0])

        for x in data:
            toinsert = {
                        'STO': x[2],
                        'GPON' : x[3],
                        'SLOT': x[4],
                        'PORT' : x[5],
                        'INET': x[8],
                        'ODP' : x[10],
                        'VALIDITAS' : x[11],
                        }
            operations.append(InsertOne(toinsert))
            pos +=1
        self.coll.drop()
        result = self.coll.bulk_write(operations)
        return pos

    def get_odp(self):
        data = self.coll.find()
        return [each['ODP'] for each in data]

    def get_inet(self):
        data = self.coll.find()
        return [each['INET'] for each in data]

    def get_data_odp(self, odp):
        data = self.coll.find({'ODP':{'$regex':odp}})
        return [[x['STO'],x['GPON'],x['SLOT'],x['PORT'],x['INET'], x['ODP'], x['VALIDITAS']] for x in data]
    
    def get_data_inet(self, inet):
        data = self.coll.find({'INET':inet})
        return [[x['STO'],x['GPON'],x['SLOT'],x['PORT'],x['INET'], x['ODP'], x['VALIDITAS']] for x in data]

class OdpLap():
    def __init__(self):
        client = MongoClient()
        self.db = client.daman
        self.coll = self.db.odpuim

    def get_started(self):
        
        operations = []
        data = nestdict()
        pos = 0
        path='D:\daman\DALAPA_VALIDASI_LAPANGAN.xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb['Sheet2']    
        data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, max_col=15, min_row=2)]
        print(data[0])

        for x in data:
            toinsert = {
                        'STO': x[2],
                        'GPON' : x[3],
                        'SLOT': x[4],
                        'PORT' : x[5],
                        'INET': x[8],
                        'ODP' : x[10],
                        'VALIDITAS' : x[11],
                        }
            operations.append(InsertOne(toinsert))
            pos +=1
        self.coll.drop()
        result = self.coll.bulk_write(operations)
        return pos

    def get_odp(self):
        data = self.coll.find()
        return [each['ODP'] for each in data]

    def get_inet(self):
        data = self.coll.find()
        return [each['INET'] for each in data]

    def get_data_odp(self, odp):
        data = self.coll.find({'ODP':{'$regex':odp}})
        return [[x['STO'],x['GPON'],x['SLOT'],x['PORT'],x['INET'], x['ODP'], x['VALIDITAS']] for x in data]
    
    def get_data_inet(self, inet):
        data = self.coll.find({'INET':inet})
        return [[x['STO'],x['GPON'],x['SLOT'],x['PORT'],x['INET'], x['ODP'], x['VALIDITAS']] for x in data]
    
    def __init__(self):
        self.data = []

    def read_data(self):
        path='D:\daman\DALAPA_VALIDASI_LAPANGAN.xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb['Sheet2']    
        self.data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=1, max_col=15, min_row=2)]
        return self.data
        
    def print_data(self, odplap):
        result = []
        if any(odplap in sublist for sublist in self.data):
            for i, x in enumerate(self.data):
              if odplap in x:
                  result.append((self.data[i][2:6]+self.data[i][8:12]))
            return result
        else:
            return False

class Absen():
    def __init__(self):
        self.data = []
        self.tanggal = []

    def read_data(self):
        mydate = datetime.datetime.now()
        currentMonth = mydate.strftime("%B")
        currentYear = mydate.strftime("%Y")
        currentSheet = currentMonth[0:3]+"-"+currentYear[2:]
        path='D:\daman\JADWAL DESEMBER 2019.xlsx'
        wb = load_workbook(path, data_only=True)
        ws = wb[currentSheet]    
        self.data = [[cell.value for cell in each] for each in ws.iter_rows(min_col=4, max_col=35, min_row=4, max_row=18)]
        self.tanggal = self.data[0][1:]
        return self.data,self.tanggal
        
    def print_data(self,tgl):
        result = []
        data = self.data[1:]
        if tgl in self.tanggal:
          index = self.tanggal.index(tgl)+1
          for i, x in enumerate(data):
            result.append(data[i][0])
            result.append(data[i][index])
        return result
